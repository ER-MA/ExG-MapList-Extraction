from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import os


# 将获取的json格式“实时地图状态列表”转换为excel表格（输出到\output文件夹中）
def save_to_excel(data, filename):
    # 创建一个工作簿
    wb = Workbook()
    # 选择当前活跃的工作表
    ws = wb.active
    # 设置工作表标题
    ws.title = "地图数据"

    # 设置表头
    headers = ["地图名", "中文名", "难度", "冷却时长", "冷却截止", "Workshop"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # 加粗表头

    # 填充数据
    json_data = json.loads(data)  # 使用json模块来解析，确保解析后的数据是字典格式
    for row_num, entry in enumerate(json_data, start=2):
        ws.cell(row=row_num, column=1, value=entry["Name"])
        ws.cell(row=row_num, column=2, value=entry["CnName"])
        ws.cell(row=row_num, column=3, value=entry["Difficulty"])
        ws.cell(row=row_num, column=4, value=entry["CooldownMinute"])
        ws.cell(row=row_num, column=5, value=entry["CooldownEnd"])
        # 创建超链接
        workshop_link = f"https://steamcommunity.com/sharedfiles/filedetails/?id={entry['Steam']}"
        ws.cell(row=row_num, column=6, value=entry['Steam'])
        # 设置超链接
        ws.cell(row=row_num, column=6).hyperlink = workshop_link
        # 设置超链接的字体颜色为蓝色并加下划线
        ws.cell(row=row_num, column=6).font = Font(color="0000FF", underline="single")

    # 调整列宽以适应内容
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 5

    # 若不存在则创建output目录，用于保存表格
    output_dir = 'output'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 保存工作簿
    current_time = datetime.now().strftime("%Y%m%d%H%M%S")  # 获取当前时间
    filename_with_timestamp = f"{filename}_({current_time}).xlsx"  # 添加时间戳及扩展名
    output_filename = os.path.join(output_dir, filename_with_timestamp)  # 构造完整的文件路径
    wb.save(output_filename)

