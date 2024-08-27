from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import os


# 将获取的json格式“实时地图状态列表”转换为excel表格（输出到\output文件夹中）
def json_to_excel(data, filename):

    # 确保传入的数据是字典格式
    if not isinstance(data, dict):
        try:
            data = json.loads(data)
        except json.JSONDecodeError:
            raise ValueError("传入的数据不是有效的JSON格式")

    # 创建一个工作簿
    wb = Workbook()
    # 选择当前活跃的工作表
    ws = wb.active
    # 设置工作表标题
    ws.title = "地图数据"

    # 设置表头
    headers = ["地图名", "中文名", "难度", "冷却时长", "冷却截止", "Workshop"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)  # 加粗表头

    # 填充数据
    for row_num, entry in enumerate(data, start=2):
        ws.cell(row=row_num, column=1, value=entry["Name"])
        ws.cell(row=row_num, column=2, value=entry["CnName"])
        ws.cell(row=row_num, column=3, value=entry["Difficulty"])
        ws.cell(row=row_num, column=4, value=entry["CooldownMinute"])
        ws.cell(row=row_num, column=5, value=entry["CooldownEnd"])
        # 创建超链接
        workshop_link = f"https://steamcommunity.com/sharedfiles/filedetails/?id={entry['Steam']}"
        # 设置超链接
        ws.cell(row=row_num, column=6, value=entry['Steam']).hyperlink = workshop_link
        # 设置超链接的字体颜色为蓝色并加下划线
        ws.cell(row=row_num, column=6).font = Font(color="0000FF", underline="single")

    # 调整列宽以适应内容
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 5

    # 若不存在则创建output目录，用于保存表格
    output_dir = 'output'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 保存工作簿
    current_time = datetime.now().strftime("%Y%m%d%H%M%S")  # 获取当前时间
    filename_with_timestamp = f"{filename}_({current_time}).xlsx"  # 添加时间戳及扩展名
    output_filename = os.path.join(output_dir, filename_with_timestamp)  # 构造完整的文件路径
    wb.save(output_filename)

    return output_filename  # 返回文件路径，以便于后续使用


def json_to_file(data, filename):

    # 确保传入的数据是字典格式
    if not isinstance(data, dict):
        try:
            data = json.loads(data)
        except json.JSONDecodeError:
            raise ValueError("传入的数据不是有效的JSON格式")

    # 格式化JSON数据
    formatted_data = json.dumps(data, indent=4, ensure_ascii=False)

    # 若不存在则创建output目录，用于保存json文件
    output_dir = 'output'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 构造完整的文件路径
    output_filename = os.path.join(output_dir, f"{filename}.json")

    # 将格式化的JSON数据写入文件
    with open(output_filename, 'w', encoding='utf-8') as file:
        file.write(formatted_data)

    return output_filename  # 返回文件路径，以便于后续使用

